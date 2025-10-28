import os
import openpyxl

def find_index(arr, value):  # 返回找到的位置
    for i in range(len(arr)):
        if arr[i] == value:
            return i
    return -1


def rewrite_lists(lists, n):
    result = []
    for coord in lists:
        result.append((n, coord))
    return result


def macro_evacuation_network_modeling(dirs_path):
    # 获取所有在该文件夹内的单层文件夹名称列表
    dirs_names = [f for f in os.listdir(dirs_path)]
    if 'Output.xlsx' in dirs_names:
        dirs_names.remove('Output.xlsx')
    if 'Output.xls' in dirs_names:
        dirs_names.remove('Output.xls')
    if 'result.csv' in dirs_names:
        dirs_names.remove('result.csv')
    if 'stairs.txt' in dirs_names:
        dirs_names.remove('stairs.txt')
    if 'min_path.txt' in dirs_names:
        dirs_names.remove('min_path.txt')
    if 'process.mp4' in dirs_names:
        dirs_names.remove('process.mp4')
    if 'multi_result.png' in dirs_names:
        dirs_names.remove('multi_result.png')

    """建立变量存储源节点、中间节点、楼梯间门节点、楼梯节点、出口节点的个数以编号"""
    nodes_list_xy_rooms = []
    nodes_list_middle_doors = []
    nodes_list_stair_points = []
    nodes_list_exit = []

    """获取多层所有节点的序列"""
    for dir_num in range(len(dirs_names)):
        fold_dir = dirs_path + '/' + str(dirs_names[dir_num])  # 单层总的文件夹 + 不加斜杠/-=

        input_path = {
            'stair_points': dirs_path + '/stairs.txt',
            'origin': fold_dir + '/2opening/seg.png',
            'draw': fold_dir + '/rgroute.png',
            'xy_doors': fold_dir + '/3approxPOLY/xy_doors.txt',
            'xy_rooms': fold_dir + '/3approxPOLY/xy_rooms.txt',
            'exit_doors': fold_dir + '/6connection/exit_doors.txt',
            'path_points': fold_dir + '/4route/path_points.txt',
            'middle_doors': fold_dir + '/4route/middle_doors.txt',
            'normal_vector': fold_dir + '/5width/normal_vector.txt',
            'width': fold_dir + '/5width/width.txt',
        }
        #  读取门、房间、出口、中间节点坐标
        xy_rooms = []
        exit_doors = []
        middle_doors = []
        f = open(input_path['xy_rooms'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            xy_rooms.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一行
        # print('xy_rooms:', xy_rooms, '\n')
        f.close()

        f = open(input_path['middle_doors'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            middle_doors.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一行
        # print('middle_doors:', middle_doors, '\n')
        f.close()

        with open(input_path['stair_points'], "r") as file:
            lines = file.readlines()
            stair_points = []
            for line in lines:
                stair_point = line.strip()[1:-1].split('), (')
                stair_point = [tuple(map(int, n.split(','))) for n in stair_point]
                stair_points.append(stair_point)

        f = open(input_path['exit_doors'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            exit_doors.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一行
        # print('exit_doors:', exit_doors, '\n')
        f.close()

        temp1 = rewrite_lists(xy_rooms, dir_num)
        temp2 = rewrite_lists(middle_doors, dir_num)
        temp4 = rewrite_lists(stair_points[dir_num], dir_num)
        temp5 = rewrite_lists(exit_doors, dir_num)
        # 坐标形式（dir_num， coordinates）

        nodes_list_xy_rooms += temp1
        nodes_list_middle_doors += temp2
        nodes_list_stair_points += temp4
        if dir_num == 0:
            nodes_list_exit += temp5

    nodes_list = []
    nodes_list = nodes_list_xy_rooms + nodes_list_middle_doors + nodes_list_stair_points + nodes_list_exit  # 创建数组用于索引节点编号
    # n = find_index(nodes_list, (层号， (381, 316)))

    """为总的工作表定义每一列的值的存储地址"""
    sheet1_column1 = []  # 第一个工作表  5列
    sheet1_column2 = []
    sheet1_column3 = []
    sheet1_column4 = []
    sheet1_column5 = []
    sheet2_column1 = []  # 第二个工作表  2列
    sheet2_column2 = []
    sheet3_column1 = []  # 第二个工作表  3列
    sheet3_column2 = []
    sheet3_column3 = []
    sheet4_column1 = []  # 第三个工作表  3列
    sheet4_column2 = []
    sheet4_column3 = []

    for dir_num in range(len(dirs_names)):
        fold_dir = dirs_path + '/' + str(dirs_names[dir_num])  # 单层总的文件夹 + 不加斜杠/-=

        input_path = {
            'xy_doors': fold_dir + '/3approxPOLY/xy_doors.txt',
            'xy_rooms': fold_dir + '/3approxPOLY/xy_rooms.txt',
            'exit_doors': fold_dir + '/6connection/exit_doors.txt',
            'path_points': fold_dir + '/4route/path_points.txt',
            'path_length': fold_dir + '/4route/path_length.txt',
            'middle_doors': fold_dir + '/4route/middle_doors.txt',
            'width': fold_dir + '/5width/width.txt',
            'ops': fold_dir + '/5width/ops.txt',
            'factor': fold_dir + '/factor.txt',
            'room_areas': fold_dir + '/3approxPOLY/room_areas.txt',
            'excel': fold_dir + '/output.xlsx',
        }

        #  读取门、房间、出口、中间节点坐标、以及比例
        xy_doors = []
        xy_rooms = []
        exit_doors = []
        middle_doors = []
        width = []
        path_length = []
        ops = []
        room_areas = []

        with open(input_path['factor'], 'r') as f:
            factor = float(f.readline())
            factor = round(factor, 4)
        print('factor:', factor)

        f = open(input_path['xy_doors'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            xy_doors.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一
        print('xy_doors:', xy_doors)  # 行

        f = open(input_path['xy_rooms'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            xy_rooms.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一行
        print('xy_rooms:', xy_rooms)

        f = open(input_path['exit_doors'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            exit_doors.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一行
        print('exit_doors:', exit_doors)

        f = open(input_path['middle_doors'], "r", encoding='utf-8')
        line = f.readline()  # 读取第一行
        while line:
            txt_data = eval(line)  # 可将字符串变为元组
            middle_doors.append(txt_data)  # 列表增加
            line = f.readline()  # 读取下一行
        print('middle_doors:', middle_doors)
        f.close()

        # 读取宽度
        with open(input_path['width'], "r") as file:
            for line in file.readlines():
                width.append(int(line.strip()))
        print('width:', width)

        # 读取长度
        with open(input_path['path_length'], 'r') as f:
            for line in f.readlines():
                t_lengths = line.strip().split(',')
                numbers = [float(t_length) for t_length in t_lengths]
                path_length.append(numbers)
        print('path_length:', path_length)

        # 读取OPS
        with open(input_path['ops'], "r") as file:
            for line in file.readlines():
                ops.append(float(line.strip()))
        print('ops:', ops)

        # 读取房间面积
        with open(input_path['room_areas'], 'r') as file:
            for line in file.readlines():
                room_areas.append(int(line.strip()))
        room_areas = [round(x * (factor ** 2), 2) for x in room_areas]
        print('room_areas:', room_areas)

        # 读取路线数据
        with open(input_path['path_points'], 'r') as f:
            lines = f.readlines()  # 读取所有行
        path_points = []  # 定义一个空列表存储读取的元组
        for line in lines:  # 遍历每一行
            points = []
            line_list = line.strip().split(')(')  # 将字符串按照括号切割
            for item in line_list:  # 遍历每个元素
                # 将括号和逗号去掉，再将每两个数字组成一个元组，添加到points列表中
                points.append(tuple(map(int, item.replace('(', '').replace(')', '').replace(',', ' ').split())))
            path_points.append(points)
        print('path_points:', path_points)  # 输出读取到的元组列表

        """将数据按格式要求输出到excel文档"""
        workbook = openpyxl.Workbook()  # 创建一个新的Excel文档
        worksheet1 = workbook.active  # 选中第一个工作表
        worksheet1.title = "Link Information"  # 修改工作表名称
        worksheet1.append(["Link ID", "Start Node", "End Node", "Length(m)", "PwT"])  # 插入表头
        """第一个工作表link Information"""
        # 遍历数据转换成所需的四个列向量
        column1 = []  # 编号
        column2 = []  # start node
        column3 = []  # end node
        column4 = []  # length
        column5 = []  # 水平为0 竖向为1
        for group in path_points:  # 写入Link Information
            for i, tpl in enumerate(group):
                if i == len(group) - 1:
                    break
                else:
                    a = find_index(nodes_list, (dir_num, group[i]))
                    b = find_index(nodes_list, (dir_num, group[i + 1]))
                    column2.append(a)
                    column3.append(b)
                    column5.append(0)
        i = 0  # 删除重复的节点间线路

        for sub_lst in path_length:  # 生成第四列
            column4.extend(sub_lst)
        column4 = [x * factor for x in column4]

        # 删除重复的节点间线路
        while i < len(column2):
            j = i + 1
            while j < len(column2):
                if column2[i] == column2[j] and column3[i] == column3[j]:
                    del column2[j]
                    del column3[j]
                    del column4[j]
                    del column5[j]
                    # 注意此处不需要将j加一，因为下一个元素会顶上来
                else:
                    j += 1
            i += 1

        # 写入
        column1 = [x for x in range(len(column2))]  # 生成第一列
        columns = [column1, column2, column3, column4, column5]
        sheet1_column2 += column2
        sheet1_column3 += column3
        sheet1_column4 += column4
        sheet1_column5 += column5

        for i in range(len(columns)):  # 写入表格
            for j in range(len(columns[i])):
                worksheet1.cell(row=j + 2, column=i + 1, value=columns[i][j])

        """第二个工作表Number of Source Nodes"""
        worksheet2 = workbook.create_sheet("Number of Source Nodes")
        worksheet2.append(["Source", "Number"])  # 插入表头

        column1 = []  # source
        column2 = []  # number

        for i in range(len(xy_rooms)):
            column1.append(find_index(nodes_list, (dir_num, xy_rooms[i])))
        for i in range(len(room_areas)):
            column2.append(int((room_areas[i] / 5) + 0.55))  # 按照实际面积÷8（以8平方米容纳一人计算）初步确定人数，四舍五入

        columns = [column1, column2]
        sheet2_column1 += column1
        sheet2_column2 += column2

        for i in range(len(columns)):  # 写入表格
            for j in range(len(columns[i])):
                worksheet2.cell(row=j + 2, column=i + 1, value=columns[i][j])

        """第三个工作表Middle Node capacity"""
        worksheet3 = workbook.create_sheet("Middle Node capacity")
        worksheet3.append(["Middle Node", "width", "ops"])  # 插入表头

        column1 = []  # middle node
        column2 = []  # width
        column3 = []  # ops

        middle_nodes = middle_doors
        for i in range(len(middle_nodes)):
            column1.append(find_index(nodes_list, (dir_num, middle_nodes[i])))  # 节点编号
            index = find_index(xy_doors, (dir_num, middle_nodes[i]))  # 门宽度索引值
            column2.append(width[index])
            column3.append(ops[index])
        column2 = [x * factor for x in column2]

        columns = [column1, column2, column3]
        sheet3_column1 += column1
        sheet3_column2 += column2
        sheet3_column3 += column3

        for i in range(len(columns)):  # 写入表格
            for j in range(len(columns[i])):
                worksheet3.cell(row=j + 2, column=i + 1, value=columns[i][j])

        """第四个工作表Exit capacity"""
        worksheet4 = workbook.create_sheet("Exit capacity")
        worksheet4.append(["exit", "width", "ops"])  # 插入表头

        column1 = []  # exit
        column2 = []  # width
        column3 = []  # ops

        if dir_num == 0:
            for i in range(len(exit_doors)):
                column1.append(find_index(nodes_list, (dir_num, exit_doors[i])))  # 节点编号
                index = find_index(xy_doors, exit_doors[i])  # 门宽度索引值
                column2.append(width[index])
                column3.append(ops[index])
        column2 = [x * factor for x in column2]
        columns = [column1, column2, column3]
        sheet4_column1 += column1
        sheet4_column2 += column2
        sheet4_column3 += column3

        for i in range(len(columns)):  # 写入表格
            for j in range(len(columns[i])):
                worksheet4.cell(row=j + 2, column=i + 1, value=columns[i][j])

        # 保存Excel文档
        workbook.save(input_path['excel'])

    """将楼梯竖向路线信息添加到总的工作表中"""
    input_path = {
        'stair_points': dirs_path + '/stairs.txt',
        'stairs_length': dirs_path + '/' + str(dirs_names[0]) + '/stairs_length.txt',
        'stairs_width': dirs_path + '/' + str(dirs_names[0]) + '/stairs_width.txt',
        'exit_doors': dirs_path + '/' + str(dirs_names[0]) + '/6connection/exit_doors.txt',
    }
    stairs_width = []
    stairs_length = []
    exit_doors=[]
    with open(input_path['stair_points'], "r") as file:
        lines = file.readlines()
        stair_points = []
        for line in lines:
            stair_point = line.strip()[1:-1].split('), (')
            stair_point = [tuple(map(int, n.split(','))) for n in stair_point]
            stair_points.append(stair_point)
    with open(input_path['stairs_width'], "r") as file:
        for line in file.readlines():
            stairs_width.append(float(line.strip()))
    print('stairs_width:', stairs_width)
    with open(input_path['stairs_length'], "r") as file:
        for line in file.readlines():
            stairs_length.append(float(line.strip()))
    print('stairs_length:', stairs_length)
    f = open(input_path['exit_doors'], "r", encoding='utf-8')
    line = f.readline()  # 读取第一行
    while line:
        txt_data = eval(line)  # 可将字符串变为元组
        exit_doors.append(txt_data)  # 列表增加
        line = f.readline()  # 读取下一行
    # print('exit_doors:', exit_doors, '\n')
    f.close()

    #  第一个工作表link Information
    #  遍历数据转换成所需的四个列向量
    column1 = []  # 编号
    column2 = []  # start node
    column3 = []  # end node
    column4 = []  # length
    column5 = []  # 水平为0 竖向为1
    for dir_num in range(len(dirs_names) - 1):
        for i in range(len(stair_points[dir_num])):  # 写入Link Information
            t_point0 = (dir_num, stair_points[dir_num][i])
            t_point1 = (dir_num + 1, stair_points[dir_num + 1][i])
            a = find_index(nodes_list, t_point0)
            b = find_index(nodes_list, t_point1)
            column2.append(b)
            column3.append(a)
            if t_point0 in nodes_list_stair_points and t_point1 in nodes_list_stair_points:
                column5.append(1)
            else:
                column5.append(0)
    for q in stairs_length:  # 生成第四列
        column4.append(q)

    # 添加出口节点到出口判断的虚拟节点的路径（距离1）
    out_node_index = len(nodes_list)
    for i in range(len(exit_doors)):
        column2.append(find_index(nodes_list, (0, exit_doors[i])))  # 节点编号
        column3.append(out_node_index)
        column4.append(1)
        column5.append(0)

    sheet1_column2 += column2
    sheet1_column3 += column3
    sheet1_column4 += column4
    sheet1_column5 += column5
    #  第三个工作表Middle Node capacity
    column1 = []  # middle node
    column2 = []  # width
    column3 = []  # ops
    for dir_num in range(len(dirs_names)):
        for i in range(len(stair_points[dir_num])):  # 写入Link Information
            t_point0 = (dir_num, stair_points[dir_num][i])
            column1.append(find_index(nodes_list, t_point0))  # 节点编号
            column2.append(stairs_width[i])
            column3.append(0.3)
    sheet3_column1 += column1
    sheet3_column2 += column2
    sheet3_column3 += column3

    """保存总的多层文件到起始目录"""
    workbook = openpyxl.Workbook()  # 创建一个新的Excel文档
    """第一个工作表link Information"""
    worksheet1 = workbook.active  # 选中第一个工作表
    worksheet1.title = "Link Information"  # 修改工作表名称
    worksheet1.append(["Link ID", "Start Node", "End Node", "Length(m)", "PwT"])  # 插入表头
    sheet1_column1 = [x for x in range(len(sheet1_column2))]
    columns = [sheet1_column1, sheet1_column2, sheet1_column3, sheet1_column4, sheet1_column5]
    for i in range(len(columns)):  # 写入表格
        for j in range(len(columns[i])):
            worksheet1.cell(row=j + 2, column=i + 1, value=columns[i][j])
    """第二个工作表Number of Source Nodes"""
    worksheet2 = workbook.create_sheet("Number of Source Nodes")
    worksheet2.append(["Source", "Number"])  # 插入表头
    columns = [sheet2_column1, sheet2_column2]
    for i in range(len(columns)):  # 写入表格
        for j in range(len(columns[i])):
            worksheet2.cell(row=j + 2, column=i + 1, value=columns[i][j])
    """第三个工作表Middle Node capacity"""
    worksheet3 = workbook.create_sheet("Middle Node capacity")
    worksheet3.append(["Middle Node", "width", "ops"])  # 插入表头
    columns = [sheet3_column1, sheet3_column2, sheet3_column3]
    for i in range(len(columns)):  # 写入表格
        for j in range(len(columns[i])):
            worksheet3.cell(row=j + 2, column=i + 1, value=columns[i][j])
    """第四个工作表Exit capacity"""
    worksheet4 = workbook.create_sheet("Exit capacity")
    worksheet4.append(["exit", "width", "ops"])  # 插入表头
    columns = [sheet4_column1, sheet4_column2, sheet4_column3]
    for i in range(len(columns)):  # 写入表格
        for j in range(len(columns[i])):
            worksheet4.cell(row=j + 2, column=i + 1, value=columns[i][j])
    workbook.save(dirs_path + '/Output.xlsx')  # 保存Excel文档
#
if __name__ == '__main__':
    macro_evacuation_network_modeling("C:/Users/GuYH/Desktop/test/temp")
